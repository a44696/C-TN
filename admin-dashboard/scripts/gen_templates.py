import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

out_dir = os.path.join(os.path.dirname(__file__), '..', 'public')
os.makedirs(out_dir, exist_ok=True)

header_fill = PatternFill('solid', fgColor='366092')
header_font = Font(bold=True, color='FFFFFF')
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ===== STUDENTS TEMPLATE =====
# Backend reads: values[1]=student_code, values[2]=full_name, values[3]=dob,
#                values[4]=gender, values[5]=phone_number, values[6]=class_name,
#                values[7]=email, values[8]=address, values[9]=major_name, values[10]=department_name
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'

headers = [
    'student_code', 'full_name', 'dob', 'gender',
    'phone_number', 'class_name', 'email',
    'address', 'major_name', 'department_name'
]
sample_data = [
    ['SV100', 'Nguyen Van An',   '2004-01-15', 'MALE',   '0912000001', 'CL01', 'sv100@tlu.edu.vn', 'Ha Noi', 'Cong nghe thong tin', 'Khoa CNTT'],
    ['SV101', 'Tran Thi Binh',  '2003-05-20', 'FEMALE', '0912000002', 'CL01', 'sv101@tlu.edu.vn', 'Ha Noi', 'Cong nghe thong tin', 'Khoa CNTT'],
    ['SV102', 'Le Van Cuong',   '2004-08-10', 'MALE',   '0912000003', 'CL02', 'sv102@tlu.edu.vn', 'Ha Noi', 'Cong nghe thong tin', 'Khoa CNTT'],
]
col_widths = [15, 22, 14, 10, 15, 10, 26, 15, 26, 15]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border

for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

out_path = os.path.join(out_dir, 'template_students.xlsx')
wb.save(out_path)
print(f'Created: {out_path}')

# ===== LECTURERS TEMPLATE =====
# Backend reads: values[1]=lecturer_code, values[2]=full_name, values[3]=department,
#                values[4]=phone_number, values[5]=email, values[6]=major_name, values[7]=degree
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Sheet1'

headers2 = [
    'lecturer_code', 'full_name', 'department',
    'phone_number', 'email', 'major_name', 'degree'
]
sample_data2 = [
    ['GV100', 'Nguyen Van Giang', 'Khoa CNTT', '0981000001', 'gv100@tlu.edu.vn', 'Cong nghe thong tin', 'MASTER'],
    ['GV101', 'Tran Thi Huong',  'Khoa CNTT', '0981000002', 'gv101@tlu.edu.vn', 'Mang may tinh',       'PHD'],
    ['GV102', 'Le Van Duc',      'Khoa CNTT', '0981000003', 'gv102@tlu.edu.vn', 'Lap trinh',           'BACHELOR'],
]
col_widths2 = [15, 22, 15, 15, 26, 26, 12]

for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

for row_idx, row_data in enumerate(sample_data2, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border

for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

out_path2 = os.path.join(out_dir, 'template_lecturers.xlsx')
wb2.save(out_path2)
print(f'Created: {out_path2}')
